import nltk
from nltk.corpus import PlaintextCorpusReader
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
import re
import openpyxl
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

wb = openpyxl.load_workbook("C:/Users/user/Desktop/Output Data Structure.xlsx")
ws = wb['Sheet1']  # or wb.active

#Inserting words lists and tokenizing
negative_words = open("C:/Users/user/Desktop/Negative.txt", "r", encoding="utf-8").read()
nw= word_tokenize(negative_words)

positive_words = open("C:/Users/user/Desktop/Positive.txt", "r", encoding="utf-8").read()
pw= word_tokenize(positive_words)

stop_words = open("C:/Users/user/Desktop/Stopwords.txt", "r", encoding="utf-8").read()
sw= word_tokenize(stop_words)


for i in range(169):
#Inserting input file
    i = i+1
    raw_file = open("E:/python projects/scraping/webpages content/{}.txt".format(i), "r", encoding="utf-8").read()
    input_file = word_tokenize(raw_file.upper())  #Convert text to uppercase and split to a list of words
    print(input_file)

#clean file
    punctuations = ["?",".","!", "(",")","'", "'", "`", ',', ':']
    stop_words = set(stopwords.words('english'))
    stop_words_upper = [x.upper() for x in stop_words]
    filtered_file = []
    for sn in input_file:
        if sn not in stop_words_upper:
            filtered_file.append(sn)
            if sn in punctuations:
                filtered_file.remove(sn)
    print(filtered_file)
    total_words_after_cleaning = len(filtered_file)
    print(total_words_after_cleaning)


#count number of positive words
    positive_word_count = []
    for w in filtered_file:
        if w in pw:
            positive_word_count.append(w)
    print(positive_word_count)
    positive_score = (len(positive_word_count))
    print(round(positive_score,2))

#count number of negative words
    negative_word_count = []
    for n in filtered_file:
        if n in nw:
            negative_word_count.append(n)
    print(negative_word_count)
    negative_score = (len(negative_word_count))
    print(round(negative_score,2))

    polarity_score = ((positive_score - negative_score) / (positive_score + negative_score + 0.000001))
    print(round(polarity_score,2))

    subjectivity_score = ((positive_score + negative_score)/ ((total_words_after_cleaning) + 0.000001))
    print(round(subjectivity_score,2))

#avg wordcount
    wordcounts = []
    sentences = raw_file.split('.')
    for sentence in sentences:
        words = sentence.split(' ')
        wordcounts.append(len(words))
    average_wordcount = sum(wordcounts)/len(wordcounts)


#Syllable Count Per Word
    count = 0
    complex_word_count = 0
    for x in filtered_file:
        vowels = ['A', 'E', 'I', 'O', 'U']
        for index in range(0, len(x)):
            if x[index] in vowels:
                count += 1
                if x[-2:] == "ES" or x[-2:] == "ED":
                    count -= 1
    print(count)

#complex word count
    complex_word_count = 0
    for word in filtered_file:
        vowels = ['A', 'E', 'I', 'O', 'U']
        nu = 0
        for index in range(0, len(word)):
            if word[index] in vowels:
                nu += 1
                if word[-2:] == "ES" or word[-2:] == "ED":
                    nu -= 1
        if nu > 2:
            complex_word_count += 1
    print(complex_word_count)

#avg sentence length
    print(average_wordcount)
#Percentage of Complex words
    percent_complex_word = complex_word_count/ total_words_after_cleaning
    print(percent_complex_word)
#Fog index
    fog_index = 0.4 * (average_wordcount + percent_complex_word)
    print(fog_index)


#Personal Pronouns
    txt = raw_file
    x = re.findall('I', txt)
    pronoun_count = len(x)
    x = re.findall('we', txt)
    pronoun_count = pronoun_count + len(x)
    x = re.findall('my', txt)
    pronoun_count = pronoun_count + len(x)
    x = re.findall('ours', txt)
    pronoun_count = pronoun_count + len(x)
    x = re.findall('us', txt)
    pronoun_count = pronoun_count + len(x)
    x = re.findall('US', txt)
    pronoun_count = pronoun_count - len(x)

    print(pronoun_count)

#Average Word Length
    total = 0
    for word in filtered_file:
        total += len(word)
    print(total)
    average_word_length = total/total_words_after_cleaning

    cell = ['C', 'D', 'E', 'F', 'G','H', 'I', 'J', 'K', 'L','M', 'N', 'O']
    for a in cell:
        ws['C{}'.format(i + 1)] = positive_score
        ws['D{}'.format(i + 1)] = negative_score
        ws['E{}'.format(i + 1)] = polarity_score
        ws['F{}'.format(i + 1)] = subjectivity_score
        ws['G{}'.format(i + 1)] = average_wordcount
        ws['H{}'.format(i + 1)] = percent_complex_word
        ws['I{}'.format(i + 1)] = fog_index
        ws['J{}'.format(i + 1)] = average_wordcount
        ws['K{}'.format(i + 1)] = complex_word_count
        ws['L{}'.format(i + 1)] = total_words_after_cleaning
        ws['M{}'.format(i + 1)] = count
        ws['N{}'.format(i + 1)] = pronoun_count
        ws['O{}'.format(i + 1)] = average_word_length

        wb.save("C:/Users/user/Desktop/Output Data Structure.xlsx")